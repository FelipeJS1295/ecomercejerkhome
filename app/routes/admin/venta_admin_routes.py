from fastapi import APIRouter, Request, Depends, Form, HTTPException, status
from fastapi.templating import Jinja2Templates
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, desc
from app.database.database import get_db
from app.models.ventas import Venta, EstadoPago, EstadoVenta
from datetime import datetime, date
from typing import Optional, List
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

router = APIRouter(prefix="/admin")
templates = Jinja2Templates(directory="app/templates")

# Modelo para el cambio masivo
class CambioMasivoRequest(BaseModel):
    venta_ids: List[int]
    nuevo_estado: str

@router.get("/ventas")
async def admin_ventas(
    request: Request, 
    db: Session = Depends(get_db),
    estado_venta: Optional[str] = None,
    estado_pago: Optional[str] = None,
    buscar: Optional[str] = None
):
    # Query base
    query = db.query(Venta)
    
    # Aplicar filtros
    if estado_venta and estado_venta != "todos":
        query = query.filter(Venta.estado_venta == estado_venta)
    
    if estado_pago and estado_pago != "todos":
        query = query.filter(Venta.estado_pago == estado_pago)
    
    if buscar:
        query = query.filter(
            or_(
                Venta.orden_compra.contains(buscar),
                Venta.nombre_cliente.contains(buscar),
                Venta.rut_cliente.contains(buscar),
                Venta.email.contains(buscar),
                Venta.sku.contains(buscar)
            )
        )
    
    # Ordenar por fecha más reciente
    ventas = query.order_by(desc(Venta.fecha_compra)).all()
    
    # Estadísticas rápidas
    total_ventas = db.query(Venta).count()
    ventas_pendientes = db.query(Venta).filter(Venta.estado_venta == EstadoVenta.NUEVA).count()
    ventas_entregadas = db.query(Venta).filter(Venta.estado_venta == EstadoVenta.ENTREGADA).count()
    
    # Total en dinero de ventas del mes actual
    mes_actual = datetime.now().replace(day=1)
    total_mes = db.query(func.sum(Venta.precio * Venta.cantidad)).filter(
        Venta.fecha_compra >= mes_actual
    ).scalar() or 0
    
    stats = {
        "total_ventas": total_ventas,
        "ventas_pendientes": ventas_pendientes,
        "ventas_entregadas": ventas_entregadas,
        "total_mes": float(total_mes)
    }
    
    return templates.TemplateResponse("admin/ventas.html", {
        "request": request,
        "ventas": ventas,
        "stats": stats,
        "estado_venta_filter": estado_venta,
        "estado_pago_filter": estado_pago,
        "buscar": buscar or ""
    })

@router.get("/ventas/exportar")
async def exportar_ventas(
    db: Session = Depends(get_db),
    estado_venta: Optional[str] = None,
    estado_pago: Optional[str] = None,
    buscar: Optional[str] = None
):
    # Aplicar los mismos filtros que en la vista principal
    query = db.query(Venta)
    
    if estado_venta and estado_venta != "todos":
        query = query.filter(Venta.estado_venta == estado_venta)
    
    if estado_pago and estado_pago != "todos":
        query = query.filter(Venta.estado_pago == estado_pago)
    
    if buscar:
        query = query.filter(
            or_(
                Venta.orden_compra.contains(buscar),
                Venta.nombre_cliente.contains(buscar),
                Venta.rut_cliente.contains(buscar),
                Venta.email.contains(buscar),
                Venta.sku.contains(buscar)
            )
        )
    
    ventas = query.order_by(desc(Venta.fecha_compra)).all()
    
    # Crear el archivo Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Ventas JerkHome"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        "ID", "Orden Compra", "Número Venta", "Fecha Compra", "Fecha Entrega",
        "Cliente", "RUT", "Email", "Teléfono", "Dirección", "Comuna", "Región",
        "Producto", "SKU", "Cantidad", "Precio Unitario", "Total",
        "Estado Venta", "Estado Pago", "Creado", "Actualizado"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, venta in enumerate(ventas, 2):
        data = [
            venta.id,
            venta.orden_compra,
            venta.numero_venta,
            venta.fecha_compra.strftime("%d/%m/%Y %H:%M") if venta.fecha_compra else "",
            venta.fecha_entrega.strftime("%d/%m/%Y") if venta.fecha_entrega else "",
            venta.nombre_cliente,
            venta.rut_cliente,
            venta.email,
            venta.numero_telefono,
            venta.direccion_cliente,
            venta.comuna_cliente,
            venta.region_cliente,
            venta.nombre,
            venta.sku,
            venta.cantidad,
            float(venta.precio),
            float(venta.total),
            venta.estado_venta.value,
            venta.estado_pago.value,
            venta.created_at.strftime("%d/%m/%Y %H:%M") if venta.created_at else "",
            venta.updated_at.strftime("%d/%m/%Y %H:%M") if venta.updated_at else ""
        ]
        
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, len(ventas) + 2):
            cell_value = str(worksheet[f"{column_letter}{row}"].value or "")
            max_length = max(max_length, len(cell_value))
        
        # Establecer ancho con un mínimo y máximo
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generar nombre del archivo
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ventas_jerkhome_{fecha_actual}.xlsx"
    
    # Retornar el archivo
    return StreamingResponse(
        io.BytesIO(excel_buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/ventas/{venta_id}")
async def detalle_venta(venta_id: int, request: Request, db: Session = Depends(get_db)):
    venta = db.query(Venta).filter(Venta.id == venta_id).first()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    return templates.TemplateResponse("admin/venta_detalle.html", {
        "request": request,
        "venta": venta
    })

@router.post("/ventas/{venta_id}/actualizar")
async def actualizar_venta(
    venta_id: int,
    request: Request,
    estado_venta: str = Form(...),
    estado_pago: str = Form(...),
    fecha_entrega: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    venta = db.query(Venta).filter(Venta.id == venta_id).first()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    # Actualizar estados
    venta.estado_venta = EstadoVenta(estado_venta)
    venta.estado_pago = EstadoPago(estado_pago)
    
    # Actualizar fecha de entrega si se proporcionó
    if fecha_entrega:
        try:
            venta.fecha_entrega = datetime.strptime(fecha_entrega, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    db.commit()
    return RedirectResponse(url=f"/admin/ventas/{venta_id}", status_code=status.HTTP_302_FOUND)

@router.get("/ventas/{venta_id}/eliminar")
async def eliminar_venta(venta_id: int, request: Request, db: Session = Depends(get_db)):
    venta = db.query(Venta).filter(Venta.id == venta_id).first()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    # Solo permitir eliminar ventas anuladas
    if venta.estado_venta != EstadoVenta.ANULADA:
        return RedirectResponse(
            url="/admin/ventas?error=solo_anuladas", 
            status_code=status.HTTP_302_FOUND
        )
    
    db.delete(venta)
    db.commit()
    
    return RedirectResponse(url="/admin/ventas", status_code=status.HTTP_302_FOUND)

@router.post("/ventas/cambio-masivo")
async def cambio_masivo_estados(
    cambio_request: CambioMasivoRequest,
    db: Session = Depends(get_db)
):
    try:
        # Validar el nuevo estado
        try:
            nuevo_estado = EstadoVenta(cambio_request.nuevo_estado)
        except ValueError:
            return JSONResponse(
                content={"success": False, "message": "Estado inválido"}, 
                status_code=400
            )
        
        # Obtener las ventas a actualizar
        ventas = db.query(Venta).filter(Venta.id.in_(cambio_request.venta_ids)).all()
        
        if not ventas:
            return JSONResponse(
                content={"success": False, "message": "No se encontraron ventas"}, 
                status_code=404
            )
        
        # Actualizar el estado de todas las ventas seleccionadas
        ventas_actualizadas = 0
        for venta in ventas:
            venta.estado_venta = nuevo_estado
            ventas_actualizadas += 1
        
        db.commit()
        
        return JSONResponse(content={
            "success": True, 
            "message": f"Se actualizaron {ventas_actualizadas} ventas correctamente",
            "ventas_actualizadas": ventas_actualizadas
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse(
            content={"success": False, "message": f"Error interno: {str(e)}"}, 
            status_code=500
        )